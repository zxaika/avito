"""Собрать data/category_templates.json из Excel «Категории.xlsx»."""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path.home() / "Desktop" / "Категории.xlsx"
OUTPUT = ROOT / "data" / "category_templates.json"

UI_HIDDEN = {
    "Title",
    "Description",
    "Price",
    "Address",
    "ContactPhone",
    "ImageUrls",
    "ImageNames",
    "Images",
    "Category",
    "ServiceType",
    "ServiceSubtype",
    "Id",
    "AvitoId",
    "AllowEmail",
}

SLUG_BY_COL = {
    1: "fundamentnye_raboty",
    3: "otdelka_derevyannykh_domov_ban_saun",
    5: "stroitelstvo_domov_pod_klyuch",
}


def build(xlsx_path: Path) -> dict[str, object]:
    ws = load_workbook(xlsx_path, data_only=True)["Категории"]
    categories = []
    for col, slug in SLUG_BY_COL.items():
        raw_path = str(ws.cell(1, col).value or "")
        path = " / ".join(part.strip() for part in raw_path.split(" - "))
        tags: list[str] = []
        for row in range(3, ws.max_row + 1):
            tag = ws.cell(row, col).value
            if not tag:
                continue
            tag_str = str(tag).strip()
            if tag_str in UI_HIDDEN:
                continue
            tags.append(tag_str)
        categories.append(
            {
                "slug": slug,
                "path": path,
                "path_dash": raw_path,
                "field_tags": tags,
            }
        )
    return {"categories": categories}


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        print(f"Файл не найден: {xlsx}", file=sys.stderr)
        return 1
    payload = build(xlsx)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Записано: {OUTPUT}")
    for item in payload["categories"]:
        print(f"  {item['slug']}: {len(item['field_tags'])} полей")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
