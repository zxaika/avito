"""Экспорт аналитики объявлений в Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.avito_service import ListingWithStats


HEADERS = [
    "ID объявления",
    "Категория",
    "Заголовок",
    "Цена",
    "Город",
    "URL",
    "Просмотры (тек.)",
    "Просмотры (пред.)",
    "Δ %",
    "Контакты",
    "Статус",
    "В фиде",
]

WARN_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")


def export_listings_report(rows: list[ListingWithStats], path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Объявления"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        listing = row.listing
        delta = "" if row.views_delta_pct is None else row.views_delta_pct
        status = "Просадка" if row.should_archive else "OK"
        sheet.append(
            [
                listing.item_id,
                listing.category_display,
                listing.title,
                listing.price,
                listing.city,
                listing.url,
                row.current_views,
                row.previous_views,
                delta,
                row.current_contacts,
                status,
                "Да" if row.in_local_feed else "Нет",
            ]
        )
        if row.should_archive:
            for cell in sheet[sheet.max_row]:
                cell.fill = WARN_FILL

    for column in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_len + 2, 50)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
